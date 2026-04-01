import streamlit as st
import json
import pandas as pd
import io
import os
import re
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

st.set_page_config(page_title="Simetrik Docs  | PeYa", page_icon="🛵📄", layout="wide")

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTES (Paleta Excel Claro Corporativo PeYa)
# ══════════════════════════════════════════════════════════════════════════════
C = {
    "red":    "EA0050", "white": "FFFFFF", "grey":  "F9FAFB", # Zebra súper sutil
    "dark":   "2B2B2B", "border":"E5E7EB", "blue":  "2563EB", # Gris oscuro corporativo
    "teal":   "0D9488", "amber": "D97706", "purple":"7C3AED",
    "green":  "16A34A", "slate": "475569", "rose":  "E11D48", 
}

RT_LABEL = {
    "native":                  "📥 Fuente",
    "source_union":            "🔗 Unión de Fuentes",
    "source_group":            "📊 Agrupación (Group By)",
    "reconciliation":          "⚖️ Conciliación Estándar",
    "advanced_reconciliation": "🔬 Conciliación Avanzada",
    "consolidation":           "🗂️ Consolidación",
    "resource_join":           "🔀 Join de Recursos",
    "cumulative_balance":      "📈 Balance Acumulado",
}

# Colores vivos para modo claro
RT_COLOR = {
    "native":                  "3B82F6", 
    "source_union":            "0D9488", 
    "source_group":            "D97706", 
    "reconciliation":          "EA0050", # PeYa Red
    "advanced_reconciliation": "7C3AED", 
    "consolidation":           "475569", 
    "resource_join":           "16A34A", 
    "cumulative_balance":      "16A34A", 
}

# Orden de tipos para sorting
RT_ORDER = {
    "native": 1, "source_union": 2, "source_group": 3,
    "reconciliation": 4, "advanced_reconciliation": 5,
    "consolidation": 6, "resource_join": 7, "cumulative_balance": 8,
}

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS OPENPYXL (Bordes perfectos en celdas combinadas)
# ══════════════════════════════════════════════════════════════════════════════
def mk_border():
    t = Side(border_style="thin", color=C["border"])
    return Border(left=t, right=t, top=t, bottom=t)

def sc(cell, bg=None, bold=False, color=C["dark"], size=10,
       ha='left', va='top', wrap=True):
    cell.border = mk_border()
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    cell.font = Font(name='Calibri', bold=bold, size=size, color=color)
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

def hdr(cell, text, bg=C["dark"]):
    cell.value = text
    sc(cell, bg=bg, bold=True, color=C["white"], size=10,
       ha='center', va='center', wrap=False)

def section_title(ws, row, text, bg=C["red"], cols=5):
    ws.merge_cells(f'A{row}:{chr(64+cols)}{row}')
    # Pintar todas las celdas debajo del merge para garantizar el borde derecho
    for i in range(1, cols + 1):
        c = ws.cell(row, i, text if i == 1 else "")
        sc(c, bg=bg, bold=True, color=C["white"], size=10,
           ha='left', va='center', wrap=False)
    ws.row_dimensions[row].height = 20
    return row + 1

def meta_row(ws, row, label, value, cols=5, bg_val=None):
    bg_val = bg_val or C["grey"]
    c_l = ws.cell(row, 1, label)
    sc(c_l, bg=C["slate"], bold=True, color=C["white"], size=9,
       ha='left', va='center', wrap=False)
    
    ws.merge_cells(f'B{row}:{chr(64+cols)}{row}')
    val_str = str(value) if value is not None else "—"
    # Pintar todas las celdas combinadas para garantizar bordes completos
    for i in range(2, cols + 1):
        c = ws.cell(row, i, val_str if i == 2 else "")
        sc(c, bg=bg_val, size=9, va='center', wrap=True)
        
    ws.row_dimensions[row].height = 14
    return row + 1

def row_height(n_lines, base=13):
    return max(14, n_lines * base)

# ══════════════════════════════════════════════════════════════════════════════
# PARSERS
# ══════════════════════════════════════════════════════════════════════════════
def build_maps(data):
    res_map   = {}   
    col_map   = {}   
    seg_map   = {}   
    meta_map  = {}   
    seg_usage = {}   

    for r in data.get('resources', []):
        eid  = r.get('export_id')
        res_map[eid] = r.get('name', str(eid))

        for c in (r.get('columns') or []):
            cid = c.get('export_id')
            col_map[cid] = c.get('label') or c.get('name') or str(cid)

        sg = r.get('source_group') or {}
        for c in sg.get('columns', []):
            cid = c.get('column_id')
            if cid and cid not in col_map:
                col_map[cid] = f"col_{cid}"
        for v in sg.get('values', []):
            cid = v.get('column_id')
            if cid and cid not in col_map:
                col_map[cid] = f"col_{cid}"

        adv = r.get('advanced_reconciliation') or {}
        for rg in adv.get('reconcilable_groups', []):
            for cs in rg.get('columns_selection', []):
                cid = cs.get('column_id')
                if cid and cid not in col_map:
                    col_map[cid] = f"col_{cid}"
            sc2 = rg.get('segmentation_config') or {}
            for m in sc2.get('segmentation_metadata', []):
                meta_map[m.get('export_id')] = m.get('value', '?')
            ccid = sc2.get('criteria_column_id')
            if ccid and ccid not in col_map:
                col_map[ccid] = f"col_{ccid}"

        for seg in (r.get('segments') or []):
            rules = []
            for fset in (seg.get('segment_filter_sets') or []):
                for rule in (fset.get('segment_filter_rules') or []):
                    rules.append(rule)
            seg_map[seg.get('export_id')] = {
                'name':        seg.get('name', ''),
                'resource':    r.get('name', ''),
                'resource_id': eid,
                'default':     seg.get('default_segment', False),
                'rules':       rules,
            }

    for r in data.get('resources', []):
        rname = r.get('name', '')

        recon = r.get('reconciliation') or {}
        sa = recon.get('segment_a_id')
        sb = recon.get('segment_b_id')
        pa = recon.get('segment_a_prefix', 'A')
        pb = recon.get('segment_b_prefix', 'B')
        if sa:
            seg_usage.setdefault(sa, []).append((rname, f"Conciliacion lado {pa}"))
        if sb:
            seg_usage.setdefault(sb, []).append((rname, f"Conciliacion lado {pb}"))

        adv = r.get('advanced_reconciliation') or {}
        for rg in adv.get('reconcilable_groups', []):
            sgid = rg.get('segment_id')
            if sgid:
                seg_usage.setdefault(sgid, []).append(
                    (rname, f"Conciliacion Avanzada lado {rg.get('prefix_side','?')}")
                )

        su = r.get('source_union') or {}
        for us in su.get('union_segments', []):
            sgid = us.get('segment_id')
            if sgid:
                seg_usage.setdefault(sgid, []).append((rname, "Union de Fuentes"))

    return res_map, col_map, seg_map, meta_map, seg_usage

def fmt_filter_rules(rules, col_map):
    lines = []
    for r in rules:
        col_name = col_map.get(r.get('column_id'), f"ID:{r.get('column_id')}")
        cond = r.get('condition', '')
        op   = r.get('operator', '')
        val  = r.get('value', '')
        lines.append(f"{cond} [{col_name}] {op} {val}".strip())
    return "\n".join(lines) if lines else "Sin filtros configurados"

def parse_transformation_logic(col, res_map, col_map):
    lines = []
    uniq = col.get('uniqueness')
    if uniq:
        dtype      = col.get('data_format', '')
        uniq_type  = uniq.get('type')
        order_keys = uniq.get('order_keys', [])
        part_keys  = uniq.get('partition_keys', [])

        if dtype == 'boolean':
            lines.append("TIPO: Booleano de duplicado")
        elif dtype == 'integer':
            lines.append("TIPO: Numeracion de duplicado")

        if order_keys:
            order_parts = []
            for ok in sorted(order_keys, key=lambda x: x.get('position', 0)):
                col_name  = col_map.get(ok.get('column_id'), f"ID:{ok.get('column_id')}")
                direction = "ASC" if ok.get('order_by', 1) == 1 else "DESC"
                order_parts.append(f"{col_name} {direction}")
            lines.append("ORDER BY: " + ", ".join(order_parts))

        if part_keys:
            part_names = [col_map.get(pk.get('column_id'), f"ID:{pk.get('column_id')}")
                          for pk in part_keys]
            lines.append("PARTITION BY (clave de duplicado):\n  " + "\n  ".join(part_names))

        return "\n".join(lines)

    v = col.get('v_lookup')
    if v:
        vs = v.get('v_lookup_set') or {}
        origin_id = vs.get('origin_source_id')
        origin = res_map.get(origin_id, f"ID:{origin_id}")
        rules = vs.get('rules', [])
        keys = " & ".join(
            "A." + col_map.get(r.get('column_a_id'), '?') +
            " = B." + col_map.get(r.get('column_b_id'), '?')
            for r in rules
        )
        lines.append("BUSCAR V EN: " + origin)
        if keys:
            lines.append("CLAVE MATCH: " + keys)

    parents = [t for t in (col.get('transformations') or []) if t.get('is_parent')]
    for t in parents:
        q = (t.get('query') or '').strip()
        if q and q.upper() != 'N/A':
            lines.append("FÓRMULA: " + q)

    return "\n".join(lines) if lines else "Campo directo / heredado"

def parse_std_reconciliation(recon, res_map, col_map, seg_map):
    if not recon:
        return None

    sa_id = recon.get('segment_a_id')
    sb_id = recon.get('segment_b_id')
    a_cfg = recon.get('a_source_settings') or {}
    b_cfg = recon.get('b_source_settings') or {}

    def resolve_side(cfg, seg_id, prefix):
        resource_name = res_map.get(cfg.get('resource_id'), '—')
        seg = seg_map.get(seg_id) or {}
        seg_name = seg.get('name', f"ID:{seg_id}")
        seg_rules = fmt_filter_rules(seg.get('rules', []), col_map)
        return {
            'prefix':        prefix,
            'resource_name': resource_name,
            'group_name':    seg_name,
            'group_filters': seg_rules,
            'is_trigger':    cfg.get('is_trigger', False),
        }

    sides = [
        resolve_side(a_cfg, sa_id, recon.get('segment_a_prefix', 'A')),
        resolve_side(b_cfg, sb_id, recon.get('segment_b_prefix', 'B')),
    ]

    rule_sets = []
    for rs in sorted(recon.get('reconciliation_rule_sets', []),
                     key=lambda x: x.get('position', 99)):
        rules_desc = []
        for rule in (rs.get('reconciliation_rules') or []):
            col_a = col_map.get(rule.get('column_a_id'), f"ID:{rule.get('column_a_id')}")
            col_b = col_map.get(rule.get('column_b_id'), f"ID:{rule.get('column_b_id')}")
            op    = rule.get('operator', '=')
            tol   = rule.get('tolerance', 0)
            tol_u = rule.get('tolerance_unit') or ''
            tol_s = f"  [tolerancia ±{tol} {tol_u}]" if tol else ""
            rules_desc.append(f"A.{col_a}  {op}  B.{col_b}{tol_s}")
        rule_sets.append({
            'pos':   rs.get('position', 0),
            'name':  rs.get('name', ''),
            'rules': rules_desc,
        })

    return {
        'sides':      sides,
        'is_chained': recon.get('is_chained', False),
        'rule_sets':  rule_sets,
    }

def parse_adv_reconciliation(adv, res_map, col_map, seg_map, meta_map):
    if not adv:
        return None

    groups = []
    for rg in (adv.get('reconcilable_groups') or []):
        prefix   = rg.get('prefix_side', '?')
        seg_id   = rg.get('segment_id')
        seg      = seg_map.get(seg_id) or {}
        seg_name = seg.get('name', f"ID:{seg_id}")
        resource_name = seg.get('resource', res_map.get(rg.get('resource_id'), '—'))
        seg_rules = fmt_filter_rules(seg.get('rules', []), col_map)

        sc2      = rg.get('segmentation_config') or {}
        crit_id  = sc2.get('criteria_column_id')
        crit_col = col_map.get(crit_id, f"ID:{crit_id}") if crit_id else "—"
        segments = [m.get('value', '') for m in sc2.get('segmentation_metadata', [])
                    if m.get('value')]

        groups.append({
            'prefix':        prefix,
            'resource_name': resource_name,
            'group_name':    seg_name,
            'group_filters': seg_rules,
            'crit_col':      crit_col,
            'segments':      segments,
        })

    rule_sets = []
    for rs in sorted(adv.get('reconciliation_rule_sets', []),
                     key=lambda x: x.get('position', 99)):
        rules_desc = []
        for rule in (rs.get('reconciliation_rules') or []):
            col_a = col_map.get(rule.get('column_a_id'), f"ID:{rule.get('column_a_id')}")
            col_b = col_map.get(rule.get('column_b_id'), f"ID:{rule.get('column_b_id')}")
            op    = rule.get('operator', '=')
            tol   = rule.get('tolerance', 0)
            tol_u = rule.get('tolerance_unit') or ''
            tol_s = f"  [tolerancia ±{tol} {tol_u}]" if tol else ""
            rules_desc.append(f"A.{col_a}  {op}  B.{col_b}{tol_s}")

        sweep = []
        for sw in (rs.get('sweep_sides') or []):
            p       = sw.get('prefix_side', '?')
            isr     = sw.get('input_sweep_resource') or {}
            meta_id = isr.get('segmentation_metadata_id')
            if meta_id:
                seg_val = meta_map.get(meta_id, f"ID:{meta_id}")
            else:
                seg_val = "(recurso completo sin segmentar)"
            sweep.append(f"Lado {p}: {seg_val}")

        rule_sets.append({
            'pos':        rs.get('position', 0),
            'name':       rs.get('name', ''),
            'cross_type': rs.get('cross_type', ''),
            'new_ver':    rs.get('is_new_version', False),
            'rules':      rules_desc,
            'sweep':      sweep,
        })

    return {'groups': groups, 'rule_sets': rule_sets}

def parse_segment_filters(segs, col_map):
    result = []
    for seg in (segs or []):
        rules = []
        for fset in (seg.get('segment_filter_sets') or []):
            for r in (fset.get('segment_filter_rules') or []):
                col_name = col_map.get(r.get('column_id'), f"ID:{r.get('column_id')}")
                rules.append(
                    f"{r.get('condition','')} [{col_name}] {r.get('operator','')} {r.get('value','')}".strip()
                )
        if rules:
            result.append({
                'seg_id': seg.get('export_id'),
                'name':   seg.get('name', ''),
                'rules':  rules,
            })
    return result

def parse_source_group(sg, col_map):
    if not sg:
        return [], []
    group_cols = [col_map.get(c.get('column_id'), f"ID:{c.get('column_id')}")
                  for c in sorted(sg.get('columns', []), key=lambda x: x.get('position', 0))]
    agg_vals   = [(v.get('function', '?'),
                   col_map.get(v.get('column_id'), f"ID:{v.get('column_id')}"))
                  for v in sorted(sg.get('values', []), key=lambda x: x.get('position', 0))]
    return group_cols, agg_vals

def limpiar_hoja(nombre, eid):
    clean = re.sub(r'[\\/*?:\[\]]', '', str(nombre))
    return (clean[:18] + "_" + str(eid))[:31]

def sort_key(r):
    return (RT_ORDER.get(r.get('resource_type', ''), 99), r.get('export_id', 0))

def build_relations(resources, nodes, res_map):
    all_ids = {r.get('export_id') for r in resources}
    rels = {r.get('export_id'): {"parents": [], "children": []} for r in resources}
    for n in nodes:
        t_id  = n.get('target')
        s_val = n.get('source')
        if not (t_id and s_val):
            continue
        s_list = s_val if isinstance(s_val, list) else [s_val]
        for sid in s_list:
            ext_a = "" if sid in all_ids else " ↗"
            ext_b = "" if t_id in all_ids else " ↗"
            if t_id in rels:
                rels[t_id]["parents"].append(res_map.get(sid, str(sid)) + ext_a)
            if sid in rels:
                rels[sid]["children"].append(res_map.get(t_id, str(t_id)) + ext_b)
    return rels


# ══════════════════════════════════════════════════════════════════════════════
# GENERADOR EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def generar_excel(data, selected_ids):
    all_resources             = data.get('resources', [])
    nodes                     = data.get('nodes', [])
    res_map, col_map, seg_map, meta_map, seg_usage = build_maps(data)

    seen, resources = set(), []
    for r in all_resources:
        eid = r.get('export_id')
        if eid in selected_ids and eid not in seen:
            seen.add(eid)
            resources.append(r)
    resources.sort(key=sort_key)

    rels      = build_relations(resources, nodes, res_map)
    map_hojas = {r.get('export_id'): limpiar_hoja(r.get('name', ''), r.get('export_id'))
                 for r in resources}

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        wb = writer.book

        # ── ÍNDICE ─────────────────────────────────────────────────────────────
        ws = wb.create_sheet("📚 Índice", 0)
        ws.sheet_view.showGridLines = False

        ws.merge_cells('A1:G1')
        for i in range(1, 8):
            c = ws.cell(1, i, "SIMETRIK DOCUMENTATION  ·  PeYa Finance Operations & Payments" if i == 1 else "")
            sc(c, bg=C["red"], bold=True, color=C["white"], size=13, ha='center', va='center', wrap=False)
        ws.row_dimensions[1].height = 32

        ws.merge_cells('A2:G2')
        for i in range(1, 8):
            c = ws.cell(2, i, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   Recursos documentados: {len(resources)}" if i == 1 else "")
            sc(c, bg=C["dark"], color=C["white"], size=9, ha='center', va='center', wrap=False)
        ws.row_dimensions[2].height = 15

        idx_hdrs = ["#", "ID", "NOMBRE DEL RECURSO", "TIPO",
                    "PROVIENE DE", "ALIMENTA A", "LINK 🔗"]
        for i, h in enumerate(idx_hdrs, 1):
            hdr(ws.cell(4, i, h), h, bg=C["dark"])
        ws.row_dimensions[4].height = 20
        
        ws.freeze_panes = "A5"

        for row_n, res in enumerate(resources, 5):
            eid     = res.get('export_id')
            rt      = res.get('resource_type', '')
            bg      = C["grey"] if row_n % 2 == 0 else C["white"]

            vals = [row_n - 4, eid, res.get('name', ''), RT_LABEL.get(rt, rt),
                    ", ".join(rels[eid]["parents"]) or "— origen",
                    ", ".join(rels[eid]["children"]) or "— fin de flujo"]
            for col_n, val in enumerate(vals, 1):
                c = ws.cell(row_n, col_n, val)
                sc(c, bg=bg, size=9, va='center', wrap=False)
                if col_n == 4:
                    c.font = Font(name='Calibri', bold=True, size=9,
                                  color=RT_COLOR.get(rt, C["dark"]))
                c.border = mk_border()

            lnk = ws.cell(row_n, 7, "Ver →")
            lnk.hyperlink = f"#'{map_hojas[eid]}'!A1"
            lnk.font = Font(name='Calibri', color="0D47A1", underline="single", size=9)
            lnk.border = mk_border()
            ws.row_dimensions[row_n].height = 15

        for col_n, w in enumerate([6, 11, 46, 26, 38, 38, 8], 1):
            ws.column_dimensions[chr(64+col_n)].width = w

        # ── HOJAS DE DETALLE ───────────────────────────────────────────────────
        for res in resources:
            eid  = res.get('export_id')
            rt   = res.get('resource_type', '')
            name = res.get('name', '')
            tc   = RT_COLOR.get(rt, C["dark"])
            COLS = 5

            ws = wb.create_sheet(map_hojas[eid])
            ws.sheet_view.showGridLines = False

            row = 1
            ws.merge_cells(f'A{row}:E{row}')
            for i in range(1, 6):
                c = ws.cell(row, i, RT_LABEL.get(rt, '') + "  ·  " + name if i == 1 else "")
                sc(c, bg=tc, bold=True, color=C["white"], size=12, ha='left', va='center', wrap=False)
            ws.row_dimensions[row].height = 30
            row += 1
            
            ws.freeze_panes = "A2"

            row = meta_row(ws, row, "ID Recurso",  eid,  cols=COLS)
            row = meta_row(ws, row, "Tipo",        RT_LABEL.get(rt, rt), cols=COLS)
            row = meta_row(ws, row, "Proviene de",
                           ", ".join(rels[eid]["parents"]) or "Origen", cols=COLS)
            row = meta_row(ws, row, "Alimenta a",
                           ", ".join(rels[eid]["children"]) or "Fin de flujo", cols=COLS)
            row += 1

            std = parse_std_reconciliation(res.get('reconciliation'), res_map, col_map, seg_map)
            if std:
                row = section_title(ws, row, "⚖️  REGLAS DE CONCILIACIÓN ESTÁNDAR", bg=C["red"], cols=COLS)
                row = section_title(ws, row, "  GRUPOS CONCILIABLES ACTIVOS", bg=C["rose"], cols=COLS)
                
                for col_n, h in enumerate(["LADO", "RECURSO", "GRUPO CONCILIABLE (ACTIVO)", "FILTROS DEL GRUPO"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg=C["rose"])
                hdr(ws.cell(row, 5, ""), "", bg=C["rose"])
                ws.merge_cells(f'D{row}:E{row}')
                ws.row_dimensions[row].height = 18
                row += 1

                for i, side in enumerate(std['sides']):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    trig = "  [TRIGGER]" if side['is_trigger'] else ""
                    c1 = ws.cell(row, 1, side['prefix'] + trig)
                    c2 = ws.cell(row, 2, side['resource_name'])
                    c3 = ws.cell(row, 3, side['group_name'])
                    c4 = ws.cell(row, 4, side['group_filters'])
                    c5 = ws.cell(row, 5, "")
                    ws.merge_cells(f'D{row}:E{row}')
                    n_lines = side['group_filters'].count('\n') + 1
                    for c, al in [(c1,'center'),(c2,'left'),(c3,'left'),(c4,'left'),(c5,'left')]:
                        sc(c, bg=bg, size=9, va='top', wrap=True, ha=al)
                    ws.row_dimensions[row].height = row_height(n_lines)
                    row += 1
                row += 1

                row = meta_row(ws, row, "Conciliación encadenada", "Sí" if std['is_chained'] else "No", cols=COLS)
                row += 1

                row = section_title(ws, row, "  RULE SETS DE MATCHING", bg="C62828", cols=COLS)
                for col_n, h in enumerate(["POS.", "NOMBRE DEL RULE SET", "REGLAS  (A vs B)"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg="C62828")
                hdr(ws.cell(row, 4, ""), "", bg="C62828")
                hdr(ws.cell(row, 5, ""), "", bg="C62828")
                ws.merge_cells(f'C{row}:E{row}')
                ws.row_dimensions[row].height = 18
                row += 1

                for i, rs in enumerate(std['rule_sets']):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    c1 = ws.cell(row, 1, rs['pos'])
                    c2 = ws.cell(row, 2, rs['name'])
                    c3 = ws.cell(row, 3, "\n".join(rs['rules']))
                    c4 = ws.cell(row, 4, "")
                    c5 = ws.cell(row, 5, "")
                    ws.merge_cells(f'C{row}:E{row}')
                    for c, al in [(c1,'center'),(c2,'left'),(c3,'left'),(c4,'left'),(c5,'left')]:
                        sc(c, bg=bg, size=9, va='top', wrap=True, ha=al)
                    ws.row_dimensions[row].height = row_height(len(rs['rules']))
                    row += 1
                row += 1

            adv_parsed = parse_adv_reconciliation(res.get('advanced_reconciliation'), res_map, col_map, seg_map, meta_map)
            if adv_parsed:
                row = section_title(ws, row, "🔬  REGLAS DE CONCILIACIÓN AVANZADA", bg=C["purple"], cols=COLS)
                row = section_title(ws, row, "  GRUPOS CONCILIABLES Y SEGMENTOS INTERNOS", bg="6A1B9A", cols=COLS)
                for col_n, h in enumerate(["LADO", "RECURSO", "GRUPO CONCILIABLE", "FILTROS DEL GRUPO", "SEGMENTOS INTERNOS"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg="6A1B9A")
                ws.row_dimensions[row].height = 18
                row += 1

                for i, g in enumerate(adv_parsed['groups']):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    segs_txt = "\n".join(g['segments']) if g['segments'] else "(sin segmentación interna)"
                    n_lines = max(g['group_filters'].count('\n') + 1, len(g['segments']) if g['segments'] else 1)
                    c1 = ws.cell(row, 1, g['prefix'])
                    c2 = ws.cell(row, 2, g['resource_name'])
                    c3 = ws.cell(row, 3, g['group_name'])
                    c4 = ws.cell(row, 4, g['group_filters'])
                    c5 = ws.cell(row, 5, segs_txt)
                    for c, al in [(c1,'center'),(c2,'left'),(c3,'left'),(c4,'left'),(c5,'left')]:
                        sc(c, bg=bg, size=9, va='top', wrap=True, ha=al)
                    ws.row_dimensions[row].height = row_height(n_lines)
                    row += 1
                row += 1

                row = section_title(ws, row, "  RULE SETS (SEGMENTO A vs SEGMENTO B)", bg="4A148C", cols=COLS)
                for col_n, h in enumerate(["POS.", "NOMBRE / TIPO", "REGLAS  (A vs B)", "SEGMENTO LADO A", "SEGMENTO LADO B"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg="4A148C")
                ws.row_dimensions[row].height = 18
                row += 1

                for i, rs in enumerate(adv_parsed['rule_sets']):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    name_txt = rs['name']
                    if rs['cross_type']: name_txt += "\n[" + rs['cross_type'] + "]"
                    if rs['new_ver']:    name_txt += "  ✦ new version"

                    seg_a = next((s.replace("Lado A: ", "") for s in rs['sweep'] if s.startswith("Lado A")), "—")
                    seg_b = next((s.replace("Lado B: ", "") for s in rs['sweep'] if s.startswith("Lado B")), "—")

                    c1 = ws.cell(row, 1, rs['pos'])
                    c2 = ws.cell(row, 2, name_txt)
                    c3 = ws.cell(row, 3, "\n".join(rs['rules']))
                    c4 = ws.cell(row, 4, seg_a)
                    c5 = ws.cell(row, 5, seg_b)
                    for c, al in [(c1,'center'),(c2,'left'),(c3,'left'),(c4,'left'),(c5,'left')]:
                        sc(c, bg=bg, size=9, va='top', wrap=True, ha=al)
                    n_lines = max(len(rs['rules']), 1)
                    ws.row_dimensions[row].height = row_height(n_lines)
                    row += 1
                row += 1

            sg = res.get('source_group')
            if sg:
                row = section_title(ws, row, "📊  CONFIGURACIÓN DE AGRUPACIÓN (GROUP BY)", bg=C["amber"], cols=COLS)
                group_cols, agg_vals = parse_source_group(sg, col_map)
                row = meta_row(ws, row, "GROUP BY (dimensiones)", " | ".join(group_cols) or "—", cols=COLS, bg_val="FFF3E0")
                agg_str = "  |  ".join(f"{fn}( {col} )" for fn, col in agg_vals)
                row = meta_row(ws, row, "Agregaciones (métricas)", agg_str or "—", cols=COLS, bg_val="FFF3E0")
                row = meta_row(ws, row, "Acumulativo", "Sí" if sg.get('is_accumulative') else "No", cols=COLS)
                row += 1

            su = res.get('source_union')
            if su:
                row = section_title(ws, row, "🔗  CONFIGURACIÓN DE UNIÓN DE FUENTES", bg=C["teal"], cols=COLS)
                for col_n, h in enumerate(["FUENTE", "GRUPO CONCILIABLE", "ROL", "FILTROS DEL GRUPO"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg=C["teal"])
                hdr(ws.cell(row, 5, ""), "", bg=C["teal"])
                ws.merge_cells(f'D{row}:E{row}')
                ws.row_dimensions[row].height = 18
                row += 1

                for i, us in enumerate(su.get('union_segments') or []):
                    seg_id   = us.get('segment_id')
                    seg_info = seg_map.get(seg_id) or {}
                    resource_name = seg_info.get('resource', seg_info.get('resource_name', f"ID:{seg_id}"))
                    group_name    = seg_info.get('name', f"ID:{seg_id}")
                    filters_text  = fmt_filter_rules(seg_info.get('rules', []), col_map)
                    rol = "TRIGGER · " + (us.get('trigger_type') or '') if us.get('is_trigger') else "Fuente adicional"
                    n_lines = max(filters_text.count('\n') + 1, 1)
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    c1 = ws.cell(row, 1, resource_name)
                    c2 = ws.cell(row, 2, group_name)
                    c3 = ws.cell(row, 3, rol)
                    c4 = ws.cell(row, 4, filters_text)
                    c5 = ws.cell(row, 5, "")
                    ws.merge_cells(f'D{row}:E{row}')
                    for c, al in [(c1,'left'),(c2,'left'),(c3,'center'),(c4,'left'),(c5,'left')]:
                        sc(c, bg=bg, size=9, va='top', wrap=True, ha=al)
                    ws.row_dimensions[row].height = row_height(n_lines)
                    row += 1
                row += 1

            segs_all = parse_segment_filters(res.get('segments', []), col_map)
            if segs_all:
                row = section_title(ws, row, "🔍  GRUPOS CONCILIABLES DEL RECURSO", bg=C["slate"], cols=COLS)
                for col_n, h in enumerate(["NOMBRE DEL GRUPO", "FILTROS APLICADOS"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg=C["slate"])
                hdr(ws.cell(row, 3, ""), "", bg=C["slate"])
                hdr(ws.cell(row, 4, ""), "", bg=C["slate"])
                hdr(ws.cell(row, 5, "USADO EN"), "USADO EN", bg=C["slate"])
                ws.merge_cells(f'B{row}:D{row}')
                ws.row_dimensions[row].height = 18
                row += 1
                for i, seg in enumerate(segs_all):
                    bg = C["grey"] if i % 2 == 0 else "FFFFFF"
                    usages = seg_usage.get(seg['seg_id'], [])
                    if usages:
                        usage_lines = [u[0] + " (" + u[1] + ")" for u in usages]
                        usage_text = "\n".join(usage_lines)
                    else:
                        usage_text = "Sin uso en flujo activo"
                    n_lines = max(len(seg['rules']), len(usages) if usages else 1)
                    c1 = ws.cell(row, 1, seg['name'])
                    c2 = ws.cell(row, 2, "\n".join(seg['rules']))
                    c3 = ws.cell(row, 3, "")
                    c4 = ws.cell(row, 4, "")
                    c5 = ws.cell(row, 5, usage_text)
                    ws.merge_cells(f'B{row}:D{row}')
                    for c in [c1, c2, c3, c4]:
                        sc(c, bg=bg, size=9, va='top', wrap=True)
                    sc(c5, bg=bg, size=9, va='top', wrap=True, color="16A34A" if usages else "64748B")
                    ws.row_dimensions[row].height = row_height(n_lines)
                    row += 1
                row += 1

            columns = sorted(res.get('columns') or [], key=lambda x: x.get('position', 0))
            if columns:
                row = section_title(ws, row, "📋  CONFIGURACIÓN DE COLUMNAS", bg=C["blue"], cols=COLS)
                for col_n, h in enumerate(["LABEL / NOMBRE", "TIPO DATO", "TIPO COL.", "LÓGICA · FÓRMULA · BUSCAR V"], 1):
                    hdr(ws.cell(row, col_n, h), h, bg=C["blue"])
                hdr(ws.cell(row, 5, ""), "", bg=C["blue"])
                ws.merge_cells(f'D{row}:E{row}')
                ws.row_dimensions[row].height = 18
                row += 1
                for i, col in enumerate(columns):
                    label    = col.get('label') or col.get('name', '')
                    dtype    = col.get('data_format', '')
                    col_type = (col.get('column_type') or '').replace('_', ' ').upper()
                    logic    = parse_transformation_logic(col, res_map, col_map)
                    bg       = C["grey"] if i % 2 == 0 else "FFFFFF"
                    c1 = ws.cell(row, 1, label)
                    c2 = ws.cell(row, 2, dtype)
                    c3 = ws.cell(row, 3, col_type)
                    c4 = ws.cell(row, 4, logic)
                    c5 = ws.cell(row, 5, "")
                    ws.merge_cells(f'D{row}:E{row}')
                    for c, al in [(c1,'left'),(c2,'center'),(c3,'center'),(c4,'left'),(c5,'left')]:
                        sc(c, bg=bg, size=9, va='top', wrap=True, ha=al)
                    ws.row_dimensions[row].height = row_height(logic.count('\n') + 1)
                    row += 1

            ws.column_dimensions['A'].width = 26
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 22
            ws.column_dimensions['E'].width = 36

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    output.seek(0)
    return output


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI (CSS Premium Light Mode PeYa)
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

/* ── LIGHT MODE PREMIUM PeYa ── */
html, body, [data-testid="stAppViewContainer"], [data-testid="stApp"],
.stApp, .main, section[data-testid="stSidebar"] {
    background-color: #F4F5F7 !important; /* Gris ultra claro para descanso visual */
    color: #1A1A1A !important;
}
[data-testid="stHeader"] { background-color: transparent !important; }
.block-container { background-color: #F4F5F7 !important; padding-top: 1.5rem !important; }

/* ── TIPOGRAFIA ── */
html, body, [class*="css"], .stMarkdown, .stCaption,
.stMetric, .stButton, .stDownloadButton,
div[data-testid], p, span, label, h1, h2, h3, h4 {
    font-family: 'Inter', sans-serif !important;
    color: #1A1A1A !important;
}
code, .font-mono { font-family: 'JetBrains Mono', monospace !important; color: #EA0050 !important; background: #FEF2F2 !important; }

/* ── CHECKBOXES ── */
[data-testid="stCheckbox"] label { color: #1A1A1A !important; }
[data-testid="stCheckbox"] > div { align-items: center; justify-content: center; height: 100%; margin-top: 25px;}

/* ── FILE UPLOADER ── */
div[data-testid="stFileUploader"] section {
    border: 2px dashed #E5E7EB !important;
    border-radius: 12px !important;
    background: #FFFFFF !important;
    transition: all 0.2s ease;
}
div[data-testid="stFileUploader"] section:hover {
    border-color: #EA0050 !important;
    background: #FFF1F2 !important;
}
div[data-testid="stFileUploader"] section button span { display: none !important; }
div[data-testid="stFileUploader"] section button {
    background: #FFFFFF !important;
    color: #1A1A1A !important;
    border: 1px solid #D1D5DB !important;
    border-radius: 8px !important;
    padding: 6px 16px !important;
    font-weight: 600 !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
}
div[data-testid="stFileUploader"] section button::after {
    content: 'Examinar archivos';
    font-family: 'Inter', sans-serif;
    font-size: 0.875rem;
    font-weight: 600;
    color: #1A1A1A;
}
div[data-testid="stFileUploader"] small,
div[data-testid="stFileUploader"] p { color: #6B7280 !important; }

/* ── MULTISELECT ── */
[data-testid="stMultiSelect"] > div > div {
    background: #FFFFFF !important;
    border: 1px solid #D1D5DB !important;
    color: #1A1A1A !important;
    border-radius: 8px !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
}
span[data-baseweb="tag"] { background: #F3F4F6 !important; color: #1A1A1A !important; border-radius: 6px !important; border: 1px solid #E5E7EB !important;}

/* ── PROGRESS BAR ── */
.stProgress > div > div { background: #EA0050 !important; }
.stProgress > div { background: #E5E7EB !important; }

/* ── BOTÓN PRIMARY ── */
div[data-testid="stButton"] button[kind="primary"] {
    background: #EA0050 !important;
    border: 1px solid #C0003A !important;
    font-size: 1rem !important;
    font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
    color: #ffffff !important;
    border-radius: 8px !important;
    box-shadow: 0 4px 12px rgba(234, 0, 80, 0.2) !important;
    transition: all 0.2s ease;
}
div[data-testid="stButton"] button[kind="primary"]:hover { 
    background: #C0003A !important; 
    transform: translateY(-1px);
    box-shadow: 0 6px 16px rgba(234, 0, 80, 0.3) !important;
}
div[data-testid="stDownloadButton"] button {
    font-family: 'Inter', sans-serif !important;
    background: #EA0050 !important;
    color: #ffffff !important;
    border: 1px solid #C0003A !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    box-shadow: 0 4px 12px rgba(234, 0, 80, 0.2) !important;
}

/* ── SCROLLBAR ── */
::-webkit-scrollbar { width: 8px; height: 8px; }
::-webkit-scrollbar-track { background: #F4F5F7; }
::-webkit-scrollbar-thumb { background: #D1D5DB; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #9CA3AF; }
</style>
""", unsafe_allow_html=True)

# ── HEADER ────────────────────────────────────────────────────────────────────
st.markdown("""
<div style='background: linear-gradient(135deg, #EA0050 0%, #C0003A 100%); padding:28px 32px;border-radius:12px;
    margin-bottom:32px; box-shadow: 0 8px 20px -5px rgba(234,0,80,0.3)'>
    <div style='color:#FFFFFF;font-family:Inter,sans-serif;
        font-size:1.75rem;font-weight:700;letter-spacing:-0.5px'>
        Simetrik Documentation
    </div>
    <div style='color:rgba(255,255,255,0.9);font-family:Inter,sans-serif;
        font-size:0.95rem;margin-top:6px; font-weight: 500;'>
        PedidosYa Finance Operations &amp; Payments &nbsp;·&nbsp; v2.2 · Jef
    </div>
</div>""", unsafe_allow_html=True)

# ── UPLOAD ────────────────────────────────────────────────────────────────────
up = st.file_uploader(
    "**Carga el JSON exportado desde Simetrik**",
    type=['json'],
    help="En Simetrik: Flujo → ⚙️ Configuración → Exportar JSON",
    label_visibility="visible"
)

if not up:
    st.markdown("""
    <div style='background:#FFFFFF;border:1px solid #E5E7EB;border-radius:12px;
        padding:48px 32px;text-align:center;margin-top:16px;box-shadow: 0 2px 4px rgba(0,0,0,0.02)'>
        <div style='font-size:2.5rem;margin-bottom:12px'>📂</div>
        <p style='color:#1A1A1A;font-size:1.1rem;font-weight:600;margin:0'>
            Arrastra el JSON aquí o usa el botón para seleccionarlo
        </p>
        <p style='color:#6B7280;font-size:0.9rem;margin:8px 0 0'>
            En Simetrik: Flujo → Configuracion → Exportar JSON
        </p>
    </div>""", unsafe_allow_html=True)
    st.stop()

try:
    data = json.load(up)
except Exception as e:
    st.error(f"Error al leer el JSON: {e}")
    st.stop()

all_resources = data.get('resources', [])
nodes         = data.get('nodes', [])

seen_load, resources_unique = set(), []
for r in all_resources:
    eid = r.get('export_id')
    if eid not in seen_load:
        seen_load.add(eid)
        resources_unique.append(r)

res_map, col_map, seg_map, meta_map, seg_usage = build_maps(data)
resources_unique.sort(key=sort_key)
rels_all = build_relations(resources_unique, nodes, res_map)

_type_counts = {}
for r in resources_unique:
    rt = r.get('resource_type', '')
    _type_counts[rt] = _type_counts.get(rt, 0) + 1

_total          = len(resources_unique)
_fuentes        = _type_counts.get('native', 0)
_uniones        = _type_counts.get('source_union', 0)
_agrupaciones   = _type_counts.get('source_group', 0)
_recons_std     = _type_counts.get('reconciliation', 0)
_recons_adv     = _type_counts.get('advanced_reconciliation', 0)
_recons_total   = _recons_std + _recons_adv
_nombre_display = up.name if len(up.name) <= 30 else up.name[:27] + "…"

def _metric_card(label, value, color="#EA0050", bg_color="#FFFFFF"):
    return (
        "<div style='background:" + bg_color + ";border:1px solid #E5E7EB;border-radius:12px;box-shadow:0 2px 4px rgba(0,0,0,0.04);"
        "padding:16px;text-align:center'>"
        "<div style='font-size:0.75rem;color:#6B7280;font-family:Inter,sans-serif;font-weight:600;"
        "margin-bottom:6px;text-transform:uppercase;letter-spacing:0.5px'>" + label + "</div>"
        "<div style='font-size:1.75rem;font-weight:700;color:" + color + ";font-family:Inter,sans-serif;line-height:1'>"
        + str(value) + "</div></div>"
    )

_cards_html = (
    "<div style='display:grid;grid-template-columns:repeat(7,1fr);gap:12px;margin-bottom:8px'>"
    + _metric_card("Total", _total, "#1A1A1A", "#F9FAFB")
    + _metric_card("Fuentes", _fuentes, f"#{RT_COLOR['native']}")
    + _metric_card("Uniones", _uniones, f"#{RT_COLOR['source_union']}")
    + _metric_card("Agrupaciones", _agrupaciones, f"#{RT_COLOR['source_group']}")
    + _metric_card("Conc. Std", _recons_std, f"#{RT_COLOR['reconciliation']}")
    + _metric_card("Conc. Avz", _recons_adv, f"#{RT_COLOR['advanced_reconciliation']}")
    + ("<div style='background:#FFFFFF;border:1px solid #E5E7EB;border-radius:12px;box-shadow:0 2px 4px rgba(0,0,0,0.04);"
       "padding:16px;text-align:left;display:flex;flex-direction:column;justify-content:center'>"
       "<div style='font-size:0.75rem;color:#6B7280;font-family:Inter,sans-serif;font-weight:600;"
       "margin-bottom:6px;text-transform:uppercase;letter-spacing:0.5px'>JSON cargado</div>"
       "<div style='font-size:0.85rem;font-weight:600;color:#1A1A1A;font-family:Inter,sans-serif;"
       "word-break:break-all;line-height:1.4'>" + _nombre_display + "</div></div>")
    + "</div>"
)
st.markdown(_cards_html, unsafe_allow_html=True)

st.markdown("<hr style='margin:28px 0;border-color:#E5E7EB'>", unsafe_allow_html=True)

# ── PASO 1: SELECCIÓN ─────────────────────────────────────────────────────────
st.markdown("<h3 style='margin-bottom:16px; font-weight: 700; color: #1A1A1A;'>1️⃣ &nbsp; Selecciona los recursos a documentar</h3>", unsafe_allow_html=True)

all_types = sorted({r.get('resource_type', '') for r in resources_unique},
                   key=lambda x: RT_ORDER.get(x, 99))

col_f1, col_f2 = st.columns([4, 1])
with col_f1:
    filtro_tipo = st.multiselect(
        "Filtrar por tipo de recurso",
        options=all_types,
        format_func=lambda x: RT_LABEL.get(x, x),
        default=all_types,
        label_visibility="collapsed",
        placeholder="Selecciona tipos de recurso a mostrar…"
    )

resources_visible = [r for r in resources_unique
                     if r.get('resource_type', '') in filtro_tipo]

bc1, bc2, bc3 = st.columns([1, 1, 6])
select_all   = bc1.button("✅ Todos", use_container_width=True)
deselect_all = bc2.button("☐ Ninguno", use_container_width=True)

if 'sel' not in st.session_state:
    st.session_state.sel = {r.get('export_id'): True for r in resources_unique}
if select_all:
    for r in resources_visible:
        st.session_state.sel[r.get('export_id')] = True
if deselect_all:
    for r in resources_visible:
        st.session_state.sel[r.get('export_id')] = False

st.write("")

tipo_groups: dict = {}
for r in resources_visible:
    tipo_groups.setdefault(r.get('resource_type', ''), []).append(r)

selected_ids = set()
for rt in sorted(tipo_groups.keys(), key=lambda x: RT_ORDER.get(x, 99)):
    group = tipo_groups[rt]
    color_hex  = RT_COLOR.get(rt, "475569")
    label  = RT_LABEL.get(rt, rt)

    st.markdown(
        f"<div style='display:flex;align-items:center;gap:12px;margin:36px 0 16px'>"
        f"<span style='background:#{color_hex}15;color:#{color_hex};border: 1px solid #{color_hex}40;padding:6px 14px;"
        f"border-radius:20px;font-size:0.85rem;font-weight:700;white-space:nowrap;'>"
        f"{label}</span>"
        f"<span style='color:#6B7280;font-size:0.85rem;font-weight:600'>{len(group)} recursos</span>"
        f"</div>",
        unsafe_allow_html=True
    )

    for r in group:
        eid   = r.get('export_id')
        name  = r.get('name', '')
        pars  = ", ".join(rels_all[eid]["parents"]) or "—"
        chils = ", ".join(rels_all[eid]["children"]) or "—"

        ca, cb = st.columns([0.3, 9.7])
        checked = ca.checkbox("", value=st.session_state.sel.get(eid, True), key=f"chk_{eid}")
        st.session_state.sel[eid] = checked
        
        # Tarjeta "World-Class" Modo Claro PeYa
        opacity = "1" if checked else "0.55"
        bg_color = "#FFFFFF" if checked else "#F9FAFB"
        border_color = "#EA0050" if checked else "#E5E7EB"
        border_width = "2px" if checked else "1px"
        box_shadow = "0 4px 12px rgba(0,0,0,0.06)" if checked else "none"

        cb.markdown(
            f"<div style='opacity:{opacity};padding:14px 18px;background:{bg_color};border-radius:8px;margin-bottom:8px;border:{border_width} solid {border_color};box-shadow:{box_shadow};transition:all 0.2s;'>"
            f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px'>"
            f"<span style='font-weight:700;font-size:0.95rem;color:#1A1A1A'>{name}</span>"
            f"<span style='font-size:0.75rem;color:#6B7280;font-family:JetBrains Mono,monospace;background:#F3F4F6;padding:4px 8px;border-radius:4px;border:1px solid #E5E7EB'>{eid}</span>"
            f"</div>"
            f"<div style='font-size:0.8rem;color:#6B7280;display:flex;gap:24px; font-weight: 500;'>"
            f"<span><span style='color:#9CA3AF'>⬅️</span> &nbsp;<span style='color:#4B5563'>{pars[:75]}{'…' if len(pars)>75 else ''}</span></span>"
            f"<span><span style='color:#9CA3AF'>➡️</span> &nbsp;<span style='color:#4B5563'>{chils[:75]}{'…' if len(chils)>75 else ''}</span></span>"
            f"</div>"
            f"</div>",
            unsafe_allow_html=True
        )
        if checked:
            selected_ids.add(eid)

# ── PASO 2: GENERAR ───────────────────────────────────────────────────────────
st.markdown("<hr style='margin:48px 0 24px;border-color:#E5E7EB'>", unsafe_allow_html=True)

n_sel = len(selected_ids)

if n_sel > 0:
    tipos_sel = {}
    for r in resources_unique:
        if r.get('export_id') in selected_ids:
            rt = r.get('resource_type', '')
            tipos_sel[rt] = tipos_sel.get(rt, 0) + 1

    badges_html = ""
    for rt, cnt in sorted(tipos_sel.items(), key=lambda x: RT_ORDER.get(x[0], 99)):
        color_hex = RT_COLOR.get(rt, "6B7280")
        label = RT_LABEL.get(rt, rt)
        badges_html += (
            "<span style='background:#" + color_hex + "15;color:#" + color_hex + ";border:1px solid #" + color_hex + "30;"
            "padding:4px 12px;border-radius:12px;font-size:0.8rem;"
            "font-weight:700;white-space:nowrap'>"
            + label + " (" + str(cnt) + ")"
            "</span> "
        )

    sel_label = "seleccionados" if n_sel != 1 else "seleccionado"
    resumen_html = (
        "<div style='background:#FFFFFF;border:1px solid #E5E7EB;border-radius:12px;padding:16px 20px;"
        "margin-bottom:20px;display:flex;align-items:center;gap:12px;flex-wrap:wrap;box-shadow:0 2px 8px rgba(0,0,0,0.04)'>"
        "<span style='font-weight:700;color:#1A1A1A;white-space:nowrap;font-size:1.05rem'>📋 "
        + str(n_sel) + " " + sel_label + ":</span>"
        + badges_html
        + "</div>"
    )
    st.markdown(resumen_html, unsafe_allow_html=True)
else:
    st.warning("Selecciona al menos un recurso para continuar.")
    st.stop()

nombre_dl = "skt_doc_" + os.path.splitext(up.name)[0] + "_" + datetime.now().strftime('%Y-%m-%d_%H%M') + ".xlsx"

if st.button("🚀  GENERAR EXCEL", type="primary", use_container_width=True):
    prog = st.progress(0, text="Iniciando...")
    try:
        prog.progress(15, text="Procesando recursos...")
        prog.progress(40, text="Resolviendo grupos conciliables...")
        prog.progress(65, text="Construyendo reglas de conciliacion...")
        excel_bytes = generar_excel(data, selected_ids)
        prog.progress(90, text="Aplicando estilos...")
        prog.progress(100, text="Listo.")
        st.success("✅ Excel generado con **" + str(n_sel) + "** recursos documentados.")
        st.markdown("""
<style>
@keyframes ride1{0%{left:-180px;opacity:0}6%{opacity:1}85%{opacity:1}100%{left:110vw;opacity:0}}
@keyframes ride2{0%{left:-180px;opacity:0}6%{opacity:1}85%{opacity:1}100%{left:110vw;opacity:0}}
@keyframes ride3{0%{left:-180px;opacity:0}6%{opacity:1}85%{opacity:1}100%{left:110vw;opacity:0}}
@keyframes wspin{from{transform:rotate(0)}to{transform:rotate(360deg)}}
@keyframes road-anim{from{stroke-dashoffset:0}to{stroke-dashoffset:-60}}
@keyframes popin2{0%{transform:translate(-50%,-50%) scale(0);opacity:0}65%{transform:translate(-50%,-50%) scale(1.1);opacity:1}100%{transform:translate(-50%,-50%) scale(1);opacity:1}}
@keyframes fadein2{0%{opacity:0;transform:translateX(-50%) translateY(8px)}100%{opacity:1;transform:translateX(-50%) translateY(0)}}
@keyframes overlay-fade{0%{opacity:1}78%{opacity:1}100%{opacity:0;pointer-events:none}}
.py-overlay{position:fixed;top:0;left:0;width:100vw;height:100vh;z-index:99999;background:rgba(255,255,255,0.92);animation:overlay-fade 4.4s ease .1s both;pointer-events:none;backdrop-filter:blur(6px)}
.py-road-svg{position:absolute;bottom:0;left:0;width:100%;height:32px}
.py-road-line{animation:road-anim .35s linear infinite}
.py-m1{position:absolute;bottom:32px;animation:ride1 3.4s cubic-bezier(.2,.8,.4,1) 0.0s both}
.py-m2{position:absolute;bottom:44px;animation:ride2 3.4s cubic-bezier(.2,.8,.4,1) 0.5s both}
.py-m3{position:absolute;bottom:24px;animation:ride3 3.4s cubic-bezier(.2,.8,.4,1) 1.0s both}
.py-w{transform-origin:50% 50%;animation:wspin .22s linear infinite}
.py-trail{position:absolute;right:100%;top:50%;transform:translateY(-50%);width:70px;height:4px;background:linear-gradient(90deg,transparent,#EA005044);border-radius:2px}
.py-check2{position:absolute;top:50%;left:50%;width:110px;height:110px;background:#EA0050;border-radius:50%;display:flex;align-items:center;justify-content:center;animation:popin2 .6s cubic-bezier(.175,.885,.32,1.275) .3s both;box-shadow:0 10px 30px rgba(234,0,80,0.3)}
.py-msg2{position:absolute;top:calc(50% + 85px);left:50%;font-family:Inter,system-ui,sans-serif;font-size:1.3rem;font-weight:700;color:#1A1A1A;background:#FFFFFF;padding:12px 36px;border-radius:30px;border:1px solid #E5E7EB;white-space:nowrap;animation:fadein2 .4s ease .8s both;letter-spacing:.2px;box-shadow:0 8px 24px rgba(0,0,0,0.1)}
</style>
<div class="py-overlay">
  <svg class="py-road-svg">
    <rect width="100%" height="32" fill="#F4F5F7"/>
    <line x1="0" y1="8" x2="100%" y2="8" stroke="#E5E7EB" stroke-width="1"/>
    <line class="py-road-line" x1="0" y1="16" x2="100%" y2="16" stroke="#EA0050" stroke-width="2.5" stroke-dasharray="36 24" opacity=".6"/>
    <line x1="0" y1="31" x2="100%" y2="31" stroke="#E5E7EB" stroke-width="1"/>
  </svg>
  <div class="py-m1"><div style="position:relative"><div class="py-trail"></div>
    <svg width="130" height="68" viewBox="0 0 110 58">
      <rect x="20" y="14" width="56" height="18" rx="7" fill="#EA0050"/>
      <polygon points="76,14 90,20 90,28 76,32" fill="#C0003A"/>
      <polygon points="76,14 84,10 90,14 88,14 78,14" fill="#cceeff" opacity=".75"/>
      <rect x="86" y="11" width="3" height="11" rx="1.5" fill="#1A1A1A"/>
      <ellipse cx="36" cy="13" rx="10" ry="7" fill="#111"/>
      <ellipse cx="39" cy="14" rx="4" ry="3" fill="#EA0050"/>
      <rect x="30" y="19" width="16" height="9" rx="3" fill="#111"/>
      <rect x="22" y="18" width="16" height="14" rx="2" fill="#EA0050" stroke="#fff" stroke-width="1.2"/>
      <text x="30" y="28.5" font-size="6.5" fill="white" text-anchor="middle" font-weight="800" font-family="Arial Black,Arial">PeYa</text>
      <rect x="16" y="26" width="10" height="3" rx="1.5" fill="#6B7280"/>
      <line x1="76" y1="29" x2="85" y2="42" stroke="#9CA3AF" stroke-width="2"/>
      <line x1="29" y1="30" x2="20" y2="42" stroke="#9CA3AF" stroke-width="2"/>
      <g transform="translate(85,43)"><circle r="11" fill="#F4F5F7"/><circle r="8" fill="#1A1A1A"/><g class="py-w"><line x1="0" y1="-6.5" x2="0" y2="6.5" stroke="#D1D5DB" stroke-width="1.5"/><line x1="-6.5" y1="0" x2="6.5" y2="0" stroke="#D1D5DB" stroke-width="1.5"/><line x1="-4.6" y1="-4.6" x2="4.6" y2="4.6" stroke="#9CA3AF" stroke-width="1"/><line x1="4.6" y1="-4.6" x2="-4.6" y2="4.6" stroke="#9CA3AF" stroke-width="1"/></g><circle r="3" fill="#EA0050"/></g>
      <g transform="translate(20,43)"><circle r="12" fill="#F4F5F7"/><circle r="9" fill="#1A1A1A"/><g class="py-w"><line x1="0" y1="-7" x2="0" y2="7" stroke="#D1D5DB" stroke-width="1.5"/><line x1="-7" y1="0" x2="7" y2="0" stroke="#D1D5DB" stroke-width="1.5"/><line x1="-5" y1="-5" x2="5" y2="5" stroke="#9CA3AF" stroke-width="1"/><line x1="5" y1="-5" x2="-5" y2="5" stroke="#9CA3AF" stroke-width="1"/></g><circle r="3.5" fill="#EA0050"/></g>
    </svg>
  </div></div>
  <div class="py-m2"><div style="position:relative"><div class="py-trail"></div>
    <svg width="108" height="56" viewBox="0 0 110 58">
      <rect x="20" y="14" width="56" height="18" rx="7" fill="#C0003A"/>
      <polygon points="76,14 90,20 90,28 76,32" fill="#A00030"/>
      <polygon points="76,14 84,10 90,14 88,14 78,14" fill="#cceeff" opacity=".7"/>
      <rect x="86" y="11" width="3" height="11" rx="1.5" fill="#1A1A1A"/>
      <ellipse cx="36" cy="13" rx="10" ry="7" fill="#EA0050"/>
      <ellipse cx="39" cy="14" rx="4" ry="3" fill="#fff" opacity=".6"/>
      <rect x="30" y="19" width="16" height="9" rx="3" fill="#EA0050"/>
      <rect x="22" y="18" width="16" height="14" rx="2" fill="#C0003A" stroke="#fff" stroke-width="1.2"/>
      <text x="30" y="28.5" font-size="6.5" fill="white" text-anchor="middle" font-weight="800" font-family="Arial Black,Arial">PeYa</text>
      <rect x="16" y="26" width="10" height="3" rx="1.5" fill="#6B7280"/>
      <line x1="76" y1="29" x2="85" y2="42" stroke="#9CA3AF" stroke-width="2"/>
      <line x1="29" y1="30" x2="20" y2="42" stroke="#9CA3AF" stroke-width="2"/>
      <g transform="translate(85,43)"><circle r="11" fill="#F4F5F7"/><circle r="8" fill="#1A1A1A"/><g class="py-w"><line x1="0" y1="-6.5" x2="0" y2="6.5" stroke="#D1D5DB" stroke-width="1.5"/><line x1="-6.5" y1="0" x2="6.5" y2="0" stroke="#D1D5DB" stroke-width="1.5"/></g><circle r="3" fill="#C0003A"/></g>
      <g transform="translate(20,43)"><circle r="12" fill="#F4F5F7"/><circle r="9" fill="#1A1A1A"/><g class="py-w"><line x1="0" y1="-7" x2="0" y2="7" stroke="#D1D5DB" stroke-width="1.5"/><line x1="-7" y1="0" x2="7" y2="0" stroke="#D1D5DB" stroke-width="1.5"/></g><circle r="3.5" fill="#C0003A"/></g>
    </svg>
  </div></div>
  <div class="py-m3"><div style="position:relative"><div class="py-trail"></div>
    <svg width="90" height="48" viewBox="0 0 110 58">
      <rect x="20" y="14" width="56" height="18" rx="7" fill="#EA0050"/>
      <polygon points="76,14 90,20 90,28 76,32" fill="#C0003A"/>
      <polygon points="76,14 84,10 90,14 88,14 78,14" fill="#cceeff" opacity=".7"/>
      <rect x="86" y="11" width="3" height="11" rx="1.5" fill="#1A1A1A"/>
      <ellipse cx="36" cy="13" rx="10" ry="7" fill="#222"/>
      <ellipse cx="39" cy="14" rx="4" ry="3" fill="#EA0050"/>
      <rect x="30" y="19" width="16" height="9" rx="3" fill="#222"/>
      <rect x="22" y="18" width="16" height="14" rx="2" fill="#EA0050" stroke="#fff" stroke-width="1.2"/>
      <text x="30" y="28.5" font-size="6.5" fill="white" text-anchor="middle" font-weight="800" font-family="Arial Black,Arial">PeYa</text>
      <rect x="16" y="26" width="10" height="3" rx="1.5" fill="#6B7280"/>
      <line x1="76" y1="29" x2="85" y2="42" stroke="#9CA3AF" stroke-width="2"/>
      <line x1="29" y1="30" x2="20" y2="42" stroke="#9CA3AF" stroke-width="2"/>
      <g transform="translate(85,43)"><circle r="11" fill="#F4F5F7"/><circle r="8" fill="#1A1A1A"/><g class="py-w"><line x1="0" y1="-6.5" x2="0" y2="6.5" stroke="#D1D5DB" stroke-width="1.5"/><line x1="-6.5" y1="0" x2="6.5" y2="0" stroke="#D1D5DB" stroke-width="1.5"/></g><circle r="3" fill="#EA0050"/></g>
      <g transform="translate(20,43)"><circle r="12" fill="#F4F5F7"/><circle r="9" fill="#1A1A1A"/><g class="py-w"><line x1="0" y1="-7" x2="0" y2="7" stroke="#D1D5DB" stroke-width="1.5"/><line x1="-7" y1="0" x2="7" y2="0" stroke="#D1D5DB" stroke-width="1.5"/></g><circle r="3.5" fill="#EA0050"/></g>
    </svg>
  </div></div>
  <div class="py-check2">
    <svg width="56" height="56" viewBox="0 0 36 36" fill="none">
      <polyline points="5,18 13,27 31,9" stroke="white" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/>
    </svg>
  </div>
  <div class="py-msg2">Pedido listo, ya puedes descargarlo</div>
</div>
""", unsafe_allow_html=True)
        st.download_button(
            label="📥  Descargar Excel",
            data=excel_bytes,
            file_name=nombre_dl,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    except Exception as e:
        prog.empty()
        st.error(f"Error al generar el Excel: {e}")
        import traceback
        st.code(traceback.format_exc())

st.markdown("<hr style='margin:32px 0;border-color:#E5E7EB'>", unsafe_allow_html=True)
st.caption("Simetrik Documentation · PeYa Finance Operations & Payments · v2.2 · Jef")
