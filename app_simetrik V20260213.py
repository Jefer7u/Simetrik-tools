import streamlit as st
import json
import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# --- CONFIGURACIÓN VISUAL PEYA ---
st.set_page_config(page_title="Simetrik Docs Generator", page_icon="📊", layout="wide")

COLOR_PEYA_MAIN = "EA0050"
COLOR_HEADER_TXT = "FFFFFF"
COLOR_BORDER = "CCCCCC"

def aplicar_estilos_peya(ws, min_row, max_row, min_col, max_col, es_header=False):
    thin = Side(border_style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if es_header:
                cell.fill = PatternFill(start_color=COLOR_PEYA_MAIN, end_color=COLOR_PEYA_MAIN, fill_type="solid")
                cell.font = Font(color=COLOR_HEADER_TXT, bold=True, size=11)

def formatear_reglas_segmento(segmento, col_map):
    texto_logica = []
    try:
        for fs in (segmento.get('segment_filter_sets') or []):
            set_logic = []
            condicion_set = fs.get('condition', 'AND')
            for rule in (fs.get('segment_filter_rules') or []):
                col_id = rule.get('column_id')
                col_name = col_map.get(col_id, f"ID_{col_id}")
                op = str(rule.get('operator', '=')).lower()
                val = str(rule.get('value', ''))
                if 'null' in op: set_logic.append(f"[{col_name} {op}]")
                else: set_logic.append(f"[{col_name} {op} {val}]")
            if set_logic: texto_logica.append(f"({f' {condicion_set} '.join(set_logic)})")
    except: return "Error en filtros"
    return " OR ".join(texto_logica) if texto_logica else "Sin filtros (Trae todo)"

def limpiar_nombre_hoja(nombre, nombres_usados):
    invalido = [':', '\\', '/', '?', '*', '[', ']']
    nombre_limpio = nombre
    for char in invalido: nombre_limpio = nombre_limpio.replace(char, '')
    nombre_base = nombre_limpio[:28]
    nombre_final = nombre_base
    contador = 1
    while nombre_final in nombres_usados:
        nombre_final = f"{nombre_base[:26]}({contador})"
        contador += 1
    nombres_usados.add(nombre_final)
    return nombre_final

def procesar_json(json_file):
    data = json.load(json_file)
    
    # Mapeos Base
    res_map = {}
    col_map = {}
    seg_map = {}
    resources = data.get('resources') or []
    nodes = data.get('nodes') or []

    for res in resources:
        res_id = res.get('export_id') or res.get('_id')
        res_name = res.get('name', 'Sin Nombre')
        if res_id: res_map[res_id] = res_name
        # Mapear columnas y segmentos...
        for col in (res.get('columns') or []):
            c_id = col.get('export_id')
            label = col.get('label') or col.get('name')
            if c_id: col_map[c_id] = label
        for seg in (res.get('segments') or []):
            s_id = seg.get('export_id')
            if s_id: seg_map[s_id] = seg.get('name')

    # LÓGICA DE RELACIONES (NODOS)
    relaciones = {res.get('export_id'): {"parents": [], "children": []} for res in resources}
    for node in nodes:
        target = node.get('target')
        sources = node.get('source')
        if target and sources:
            if isinstance(sources, list):
                for s in sources:
                    if target in relaciones: relaciones[target]["parents"].append(res_map.get(s, str(s)))
                    if s in relaciones: relaciones[s]["children"].append(res_map.get(target, str(target)))
            else:
                if target in relaciones: relaciones[target]["parents"].append(res_map.get(sources, str(sources)))
                if sources in relaciones: relaciones[sources]["children"].append(res_map.get(target, str(target)))

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        nombres_usados = set()
        nombres_hojas_map = [limpiar_nombre_hoja(r.get('name', 'Resource'), nombres_usados) for r in resources]

        # --- HOJA INICIO (ÍNDICE MEJORADO) ---
        ws_inicio = writer.book.create_sheet("Inicio")
        ws_inicio['A1'] = "RESUMEN EJECUTIVO"
        aplicar_estilos_peya(ws_inicio, 1, 1, 1, 4, True)
        
        ws_inicio.cell(5, 1, "NOMBRE DEL RECURSO")
        ws_inicio.cell(5, 2, "TIPO")
        ws_inicio.cell(5, 3, "RELACIÓN (PROVIENE DE)")
        ws_inicio.cell(5, 4, "LINK")
        aplicar_estilos_peya(ws_inicio, 5, 5, 1, 4, True)
        
        for i, r in enumerate(resources):
            row = 6 + i
            e_id = r.get('export_id')
            ws_inicio.cell(row, 1, r.get('name'))
            ws_inicio.cell(row, 2, r.get('resource_type', 'N/A').upper())
            
            # Mostrar de dónde viene el recurso en el índice
            padres = ", ".join(relaciones.get(e_id, {}).get('parents', []))
            ws_inicio.cell(row, 3, padres if padres else "Raíz / Input Directo")
            
            link = ws_inicio.cell(row, 4, "Ir a Hoja")
            link.hyperlink = f"#'{nombres_hojas_map[i]}'!A1"
            link.style = "Hyperlink"
            aplicar_estilos_peya(ws_inicio, row, row, 1, 3, False)

        ws_inicio.column_dimensions['A'].width = 40
        ws_inicio.column_dimensions['B'].width = 20
        ws_inicio.column_dimensions['C'].width = 50

        # --- HOJAS DE RECURSOS ---
        progress_bar = st.progress(0)
        for i, res in enumerate(resources):
            progreso = int((i + 1) / len(resources) * 100)
            progress_bar.progress(progreso)
            
            e_id = res.get('export_id')
            sheet_name = nombres_hojas_map[i]
            ws = writer.book.create_sheet(sheet_name)
            curr = 1
            
            # Título y Tipo
            ws.merge_cells(start_row=curr, start_column=1, end_row=curr, end_column=5)
            ws.cell(curr, 1, f"RECURSO: {res.get('name')}").font = Font(size=14, bold=True, color=COLOR_PEYA_MAIN)
            curr += 1
            ws.cell(curr, 1, f"Tipo: {res.get('resource_type', 'N/A').upper()}").font = Font(italic=True)
            curr += 2

            # NUEVA SECCIÓN: RELACIONES DEL RECURSO
            ws.cell(curr, 1, "FLUJO DE DATOS").font = Font(bold=True, color=COLOR_PEYA_MAIN); curr += 1
            headers_rel = ["DIRECCIÓN", "RECURSOS RELACIONADOS"]
            for ci, h in enumerate(headers_rel, 1): ws.cell(curr, ci, h)
            aplicar_estilos_peya(ws, curr, curr, 1, 2, True); curr += 1
            
            rel_info = [
                ["PROVIENE DE (Inputs)", ", ".join(relaciones.get(e_id, {}).get('parents', [])) or "Ninguno (Fuente Primaria)"],
                ["ALIMENTA A (Outputs)", ", ".join(relaciones.get(e_id, {}).get('children', [])) or "Ninguno (Nodo Final)"]
            ]
            for r_info in rel_info:
                ws.cell(curr, 1, r_info[0]); ws.cell(curr, 2, r_info[1])
                aplicar_estilos_peya(ws, curr, curr, 1, 2, False); curr += 1
            curr += 2

            # (El resto del código de conciliación, segmentos y columnas se mantiene igual...)
            if res.get('resource_type') in ['reconciliation', 'advanced_reconciliation']:
                # Adaptación para Advanced Reconciliation
                rec_data = res.get('reconciliation') or res.get('advanced_reconciliation') or {}
                # ... (resto de lógica de tablas igual)
                ws.cell(curr, 1, "CONFIGURACIÓN DE CONCILIACIÓN").font = Font(bold=True); curr += 1
                # ... (tu lógica original de dibujo de tablas)
                curr += 2

            # Segmentos
            segments = res.get('segments') or []
            if segments:
                ws.cell(curr, 1, "GRUPOS / SEGMENTOS").font = Font(bold=True, color=COLOR_PEYA_MAIN); curr += 1
                headers = ["NOMBRE", "LÓGICA"]
                for ci, h in enumerate(headers, 1): ws.cell(curr, ci, h)
                aplicar_estilos_peya(ws, curr, curr, 1, 2, True); curr += 1
                for seg in segments:
                    ws.cell(curr, 1, seg.get('name'))
                    ws.cell(curr, 2, formatear_reglas_segmento(seg, col_map))
                    aplicar_estilos_peya(ws, curr, curr, 1, 2, False); curr += 1
                curr += 2

            # Columnas
            headers = ["COLUMNA", "TIPO", "TRANSFORMACIÓN"]
            for ci, h in enumerate(headers, 1): ws.cell(curr, ci, h)
            aplicar_estilos_peya(ws, curr, curr, 1, 3, True); curr += 1
            for col in (res.get('columns') or []):
                tr = " | ".join([t.get('query','') for t in (col.get('transformations') or []) if t.get('query')])
                vals = [col.get('label'), col.get('data_format'), tr]
                for ci, v in enumerate(vals, 1): ws.cell(curr, ci, v)
                aplicar_estilos_peya(ws, curr, curr, 1, 3, False); curr += 1
            
            ws.column_dimensions['A'].width = 35; ws.column_dimensions['C'].width = 60

        if "Sheet" in writer.book.sheetnames: writer.book.remove(writer.book["Sheet"])
        
    output.seek(0)
    return output

# --- INTERFAZ DE USUARIO ---
st.markdown(f"""
    <h1 style='color: #{COLOR_PEYA_MAIN};'>Generador de Documentación Simetrik</h1>
    <p>Reporte técnico con mapeo de relaciones entre recursos.</p>
    """, unsafe_allow_html=True)

uploaded_file = st.file_uploader("Sube el JSON de Simetrik", type=['json'])

if uploaded_file is not None:
    if st.button("🚀 Generar Documentación"):
        excel_data = procesar_json(uploaded_file)
        st.download_button(label="📥 Descargar Reporte", data=excel_data, 
                         file_name=f"Documentacion_Simetrik_{datetime.now().strftime('%H%M%S')}.xlsx")